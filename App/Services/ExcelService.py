import os
import pandas as pd
from typing import Optional,Union,List
from openpyxl.utils import get_column_letter

class ExcelService(object):
    def __init__(self,rootPath:str)->None:
        self.rootPath=rootPath

    def GetFilePath(self,fileName:str)->str:
        return os.path.join(self.rootPath,"Files",fileName)

    def ReadExcel(self,fileName:str,sheetName:str,startRow:int=1,startCol:int=1,endRow:int=128,endCol:int=10)->Optional[Union[pd.DataFrame,str]]:
        filePath=self.GetFilePath(fileName)
        try:
            return pd.read_excel(filePath,sheet_name=sheetName,usecols=pd.RangeIndex(start=startCol-1,stop=endCol),skiprows=range(startRow-1),engine="openpyxl")
        except FileNotFoundError:
            return f"Error: File not found - {fileName}"
        except Exception as e:
            return f"Error while reading Excel file: {e}"
    
    def ExcelReferencetoIndex(self,referenceAddress:str):
        row=0
        col=0
        for i,char in enumerate(referenceAddress):
            if(char.isdigit()):
                row=int(referenceAddress[i:])-1
                break
            col=col*26+(ord(char.upper())-ord("A"))+1
        return row,col-1
    
    def IndextoExcelReference(self,row:int,col:int):
        colStr=""
        while col>=0:
            colStr=chr(col%26+ord("A"))+colStr
            col=col//26-1
            return f"{colStr}{row+1}"
        
    def ParseRangetoCellIndices(self,cellRange:str):
        startAddress,endAddress=cellRange.split(":")
        startRow,startCol=self.ExcelReferencetoIndex(startAddress)
        endRow,endCol=self.ExcelReferencetoIndex(endAddress)
        return startRow,startCol,endRow,endCol

    def WriteExcel(self,fileName:str,sheetName:str,data:pd.DataFrame,updateCellRange:str=None)->bool:
        path=self.GetFilePath(fileName)
        fileExists=os.path.isfile(path)
        try:
            if fileExists:
                with pd.ExcelWriter(path,engine="openpyxl",mode="a",if_sheet_exists="replace") as writer:
                    data.to_excel(writer,sheet_name=sheetName,index=False,header=False)
                return True
            else:
                with pd.ExcelWriter(path,engine="openpyxl",mode="w") as writer:
                    data.to_excel(writer,sheet_name=sheetName,index=False,header=False)
                return True
        except:
            return False
        
        # path=self.GetFilePath(fileName)
        # fileExists=os.path.isfile(path)
        # try:
        #     if fileExists:
        #         if updateCellRange:
        #             writer=pd.ExcelWriter(fileName,engine="openpyxl",mode="a",if_sheet_exists="overlay")
        #             startRow,startCol,endRow,endCol=self.ParseRangetoCellIndices(updateCellRange)
        #             for row in range(startRow-1,endRow):
        #                 for col in range(startCol-1,endCol):
        #                     sheetName.cell(row=row+1,col=col+1).value=None
        #             dataFrame.to_excel(writer,sheet_name=sheetName,startrow=startRow-1,startcol=startCol-1,index=False,header=False)
        #             writer.save()
        #         return True
        #     else:
        #         with pd.ExcelWriter(path,engine="openpyxl",mode="w") as writer:
        #             dataFrame.to_excel(writer,sheet_name=sheetName,index=False,header=False)
        #         return True
        # except:
        #     return False

    def RemoveFile(self,fileName:str)->bool:
        path=self.GetFilePath(fileName)
        if(os.path.exists(path)):
            try:
                os.remove(path)
                return True
            except Exception as e:
                print(f"Error while removing file: {e}")
                return False
        return False
    
    def GetCellAddress(self,fileName:str,sheetName:str,searchString:str,startRow:int=1,startCol:int=1,endRow:int=128,endCol:int=10)->Optional[Union[List[str],str,None]]:
        try:
            df=self.ReadExcel(fileName,sheetName,startRow,startCol,endRow,endCol)
            cellsContainingtheSearchString=[]
            for rowIndex,row in df.iterrows():
                actualRow=rowIndex+startRow+1
                for colIndex in range(startCol,endCol+1):
                    value=row.iloc[colIndex-startCol]
                    if(pd.isna(value)):
                        continue
                    if(str(value)==searchString):
                        colLetter=get_column_letter(colIndex)
                        cellAddress=f"{colLetter}{actualRow}"
                        cellsContainingtheSearchString.append(cellAddress)
            if(cellsContainingtheSearchString):
                return cellsContainingtheSearchString
            else:
                return None
        except Exception as e:
            return f"{e}"

    def ReadCell(self,fileName:str,sheetName:str,cellAddress:str)->Optional[Union[str,float]]:
        df=self.ReadExcel(fileName,sheetName)
        row,col=self.ExcelReferencetoIndex(cellAddress)
        if(row<df.shape[0] and col<df.shape[1]):
            return df.iloc[row-1,col]
        return None
    
    def ReadCellRange(self,fileName:str,sheetName:str,cellRange:str)->Union[pd.DataFrame,str]:
        dataFrame=self.ReadExcel(fileName,sheetName)
        startAddress,endAddress=cellRange.split(":")
        startRow,startCol=self.ExcelReferencetoIndex(startAddress)
        endRow,endCol=self.ExcelReferencetoIndex(endAddress)
        try:
            return dataFrame.iloc[startRow:endRow+1,startCol:endCol+1]
        except IndexError as e:
            return f"Error while reading range: {e}"
        
    def TransposeofDataFrame(self,dataFrame:pd.DataFrame)->Union[pd.DataFrame,str]:
        return dataFrame.T

    def WriteDatatoDataFrame(self,dataFrame:pd.DataFrame,cellAddress:str,data:Union[str,int,float])->Union[bool,str]:
        try:
            row,col=self.ExcelReferencetoIndex(cellAddress)
            dataFrame.iloc[row,col]=data
            return True
        except Exception as e:
            return f"Error: {e}"

    def CheckInstantCell(self,fileName:str,sheetName:str,cellAddress:str)->bool:
        df=self.ReadExcel(fileName,sheetName)
        row,col=self.ExcelReferencetoIndex(cellAddress)
        if(row<df.shape[0]):
            return pd.notna(df.iloc[row-1,col])
        return False

    def CheckNextLine(self,fileName:str,sheetName:str,cellAddress:str)->bool:
        df=self.ReadExcel(fileName,sheetName)
        row,col=self.ExcelReferencetoIndex(cellAddress)
        if(row+1<df.shape[0]):
            return pd.notna(df.iloc[row,col])
        return False

    def CheckRightCol(self,fileName:str,sheetName:str,cellAddress:str)->bool:
        df=self.ReadExcel(fileName,sheetName)
        row,col=self.ExcelReferencetoIndex(cellAddress)
        if(col+1<df.shape[1]):
            return pd.notna(df.iloc[row,col+1])
        return False